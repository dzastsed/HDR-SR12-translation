import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency 'openpyxl'. Install it with: pip install openpyxl"
    ) from exc

TABLE_START = 0x0004
DEFAULT_HEADER_WORD = 0x00001C21
ENCODING_RE = re.compile(r"decoded as\s+([A-Za-z0-9_\-]+)", re.IGNORECASE)


def choose_file_with_dialog(title: str, filetypes: List[Tuple[str, str]]) -> Optional[Path]:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()

    if not selected:
        return None
    return Path(selected)


def parse_hex_cell(value: Optional[str]) -> bytes:
    if value is None:
        return b""
    text = str(value).strip().replace(" ", "")
    if text == "":
        return b""
    return bytes.fromhex(text)


def try_decode(raw: bytes, encoding: str) -> Optional[str]:
    try:
        return raw.decode(encoding)
    except Exception:
        return None


def find_header_row(sheet) -> int:
    for row_idx in range(1, min(50, sheet.max_row) + 1):
        row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, 20)]
        normalized = [str(v).strip() if v is not None else "" for v in row_values]
        if "Index" in normalized and "DecodedText" in normalized:
            return row_idx
    raise ValueError("Could not find header row with 'Index' and 'DecodedText'.")


def get_col_index_map(sheet, header_row: int) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        key = sheet.cell(row=header_row, column=col).value
        if key is None:
            continue
        col_map[str(key).strip()] = col
    required = ["Index", "DecodedText"]
    missing = [name for name in required if name not in col_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return col_map


def extract_source_file_from_sheet(sheet) -> Optional[Path]:
    for row_idx in range(1, min(20, sheet.max_row) + 1):
        a = sheet.cell(row=row_idx, column=1).value
        b = sheet.cell(row=row_idx, column=2).value
        if str(a).strip() == "Source file" and b:
            return Path(str(b))
    return None


def pick_encoding(note_value: Optional[str], fallback: str) -> str:
    note = str(note_value) if note_value is not None else ""
    match = ENCODING_RE.search(note)
    if match:
        return match.group(1)
    return fallback


def parse_offset_cell(value: object) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    if text.lower().startswith("0x"):
        return int(text, 16)
    return int(text)


def build_rows(sheet, col_map: Dict[str, int], header_row: int) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    index_col = col_map["Index"]
    text_col = col_map["DecodedText"]
    raw_col = col_map.get("RawBytesHex")
    note_col = col_map.get("Note")
    offset_col = col_map.get("StringOffsetHex")

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        index_val = sheet.cell(row=row_idx, column=index_col).value
        text_val = sheet.cell(row=row_idx, column=text_col).value

        if index_val is None and text_val is None:
            continue

        text = "" if text_val is None else str(text_val)
        raw_hex = None if raw_col is None else sheet.cell(row=row_idx, column=raw_col).value
        note_val = None if note_col is None else sheet.cell(row=row_idx, column=note_col).value
        offset_val = None if offset_col is None else sheet.cell(row=row_idx, column=offset_col).value

        rows.append(
            {
                "index": int(index_val) if index_val is not None else len(rows),
                "text": text,
                "raw_hex": raw_hex,
                "note": note_val,
                "string_offset": parse_offset_cell(offset_val),
                "row_idx": row_idx,
            }
        )

    rows.sort(key=lambda x: int(x["index"]))
    return rows


def choose_header_prefix(
    template: Optional[Path], source_hint: Optional[Path], force_word: Optional[int]
) -> bytes:
    if force_word is not None:
        return force_word.to_bytes(4, "little")

    for candidate in [template, source_hint]:
        if candidate and candidate.exists() and candidate.is_file():
            data = candidate.read_bytes()
            if len(data) >= 4:
                return data[:4]

    return DEFAULT_HEADER_WORD.to_bytes(4, "little")


def encode_text_rows(
    rows: List[Dict[str, object]],
    default_encoding: str,
    strict: bool,
) -> Tuple[List[bytes], List[str]]:
    encoded_rows: List[bytes] = []
    warnings: List[str] = []

    for item in rows:
        text = str(item["text"])
        row_idx = int(item["row_idx"])
        row_encoding = pick_encoding(item.get("note"), default_encoding)
        raw_hex = item.get("raw_hex")

        used_raw = False
        raw_bytes = b""

        if raw_hex is not None:
            try:
                raw_bytes = parse_hex_cell(str(raw_hex))
                decoded = try_decode(raw_bytes, row_encoding)
                if decoded is not None:
                    decoded_norm = decoded.replace("\r\n", "\n")
                    text_norm = text.replace("\r\n", "\n")
                else:
                    decoded_norm = None
                    text_norm = text

                if decoded_norm is not None and decoded_norm == text_norm:
                    encoded_rows.append(raw_bytes + b"\x00")
                    used_raw = True
            except ValueError:
                warnings.append(f"Row {row_idx}: invalid RawBytesHex; encoding text instead.")

        if used_raw:
            continue

        try:
            payload = text.encode(row_encoding)
        except UnicodeEncodeError as exc:
            if strict:
                raise ValueError(
                    f"Row {row_idx}: cannot encode text with {row_encoding}: {exc}"
                ) from exc
            payload = text.encode(row_encoding, errors="replace")
            warnings.append(
                f"Row {row_idx}: unencodable chars replaced while encoding as {row_encoding}."
            )

        encoded_rows.append(payload + b"\x00")

    return encoded_rows, warnings


def encode_single_row(
    row: Dict[str, object],
    default_encoding: str,
    strict: bool,
) -> Tuple[bytes, List[str]]:
    payloads, warnings = encode_text_rows([row], default_encoding=default_encoding, strict=strict)
    return payloads[0], warnings


def find_null_terminated_length(data: bytes, start: int) -> int:
    end = start
    while end < len(data) and data[end] != 0:
        end += 1
    if end >= len(data):
        return len(data) - start
    return (end - start) + 1


def build_inplace_from_template(
    template_bytes: bytes,
    rows: List[Dict[str, object]],
    default_encoding: str,
    strict: bool,
) -> Tuple[bytes, List[str]]:
    out = bytearray(template_bytes)
    warnings: List[str] = []

    rows_by_offset: Dict[int, Dict[str, object]] = {}
    for row in rows:
        offset = row.get("string_offset")
        if not isinstance(offset, int):
            continue
        if offset in rows_by_offset:
            prev_text = str(rows_by_offset[offset].get("text", ""))
            now_text = str(row.get("text", ""))
            if prev_text != now_text:
                warnings.append(
                    f"Offset 0x{offset:08X} appears multiple times with different text; first occurrence used."
                )
            continue
        rows_by_offset[offset] = row

    valid_offsets = sorted(o for o in rows_by_offset.keys() if 0 <= o < len(out))
    if not valid_offsets:
        raise ValueError(
            "No usable StringOffsetHex values found for in-place rebuild. Use --mode repack instead."
        )

    for idx, offset in enumerate(valid_offsets):
        row = rows_by_offset[offset]
        allocated = find_null_terminated_length(template_bytes, offset)

        payload, row_warnings = encode_single_row(
            row=row,
            default_encoding=default_encoding,
            strict=strict,
        )
        warnings.extend(row_warnings)

        if len(payload) > allocated:
            message = (
                f"Row {row['row_idx']} at offset 0x{offset:08X} is too long ({len(payload)} bytes) "
                f"for allocated space ({allocated} bytes)."
            )
            if strict:
                raise ValueError(message)
            warnings.append(message + " Keeping original template bytes for this row.")
            continue

        end = offset + allocated
        out[offset:end] = b"\x00" * allocated
        out[offset:offset + len(payload)] = payload

    out_of_range_rows = [
        row for row in rows if isinstance(row.get("string_offset"), int) and row["string_offset"] >= len(out)
    ]
    if out_of_range_rows:
        warnings.append(
            f"{len(out_of_range_rows)} rows have StringOffsetHex outside template size and were ignored."
        )

    return bytes(out), warnings


def build_binary(
    encoded_rows: List[bytes],
    header_prefix: bytes,
    entry_size: int,
) -> bytes:
    if entry_size not in (4, 8):
        raise ValueError("entry_size must be 4 or 8")

    string_start = TABLE_START + (len(encoded_rows) * entry_size)
    pointer_table = bytearray()
    blob = bytearray()

    current_offset = string_start
    for payload in encoded_rows:
        pointer_table += current_offset.to_bytes(4, "little")
        if entry_size == 8:
            pointer_table += b"\x00\x00\x00\x00"
        blob += payload
        current_offset += len(payload)

    return header_prefix + bytes(pointer_table) + bytes(blob)


def build_addressed_binary(
    rows: List[Dict[str, object]],
    header_prefix: bytes,
    entry_size: int,
    default_encoding: str,
    strict: bool,
    fixed_size: Optional[int] = None,
) -> Tuple[bytes, List[str]]:
    if entry_size not in (4, 8):
        raise ValueError("entry_size must be 4 or 8")

    encoded_rows, warnings = encode_text_rows(
        rows=rows,
        default_encoding=default_encoding,
        strict=strict,
    )

    pointers: List[int] = []
    data_by_offset: Dict[int, bytes] = {}

    table_size = TABLE_START + (len(rows) * entry_size)
    max_end = table_size

    for idx, row in enumerate(rows):
        offset = row.get("string_offset")
        row_idx = int(row["row_idx"])
        payload = encoded_rows[idx]

        if not isinstance(offset, int) or offset < 0:
            message = (
                f"Row {row_idx}: invalid StringOffsetHex '{offset}'. "
                "Using 0x00000000 in pointer table."
            )
            if strict:
                raise ValueError(message)
            warnings.append(message)
            pointers.append(0)
            continue

        pointers.append(offset)

        if offset < table_size:
            message = (
                f"Row {row_idx}: StringOffsetHex 0x{offset:08X} overlaps pointer table "
                f"(table end 0x{table_size - 1:08X})."
            )
            if strict:
                raise ValueError(message)
            warnings.append(message + " Keeping pointer, skipping payload write.")
            continue

        if fixed_size is not None and offset >= fixed_size:
            warnings.append(
                f"Row {row_idx}: pointer 0x{offset:08X} is outside fixed file size; pointer kept, payload skipped."
            )
            continue

        if offset in data_by_offset:
            if data_by_offset[offset] != payload:
                warnings.append(
                    f"Row {row_idx}: duplicate offset 0x{offset:08X} with different payload; first payload kept."
                )
            continue

        if fixed_size is not None and (offset + len(payload)) > fixed_size:
            message = (
                f"Row {row_idx}: payload at 0x{offset:08X} would exceed fixed file size "
                f"({offset + len(payload)} > {fixed_size})."
            )
            if strict:
                raise ValueError(message)
            warnings.append(message + " Pointer kept, payload skipped.")
            continue

        data_by_offset[offset] = payload
        max_end = max(max_end, offset + len(payload))

    if fixed_size is not None:
        if fixed_size < table_size:
            raise ValueError(
                f"Fixed size {fixed_size} is too small for pointer table size {table_size}."
            )
        out_size = fixed_size
    else:
        out_size = max_end

    out = bytearray(out_size)
    out[0:4] = header_prefix

    for idx, pointer in enumerate(pointers):
        pos = TABLE_START + (idx * entry_size)
        out[pos:pos + 4] = pointer.to_bytes(4, "little")
        if entry_size == 8:
            out[pos + 4:pos + 8] = b"\x00\x00\x00\x00"

    for offset, payload in sorted(data_by_offset.items()):
        out[offset:offset + len(payload)] = payload

    return bytes(out), warnings


def build_output_path(xlsx_path: Path, output_override: Optional[Path]) -> Path:
    if output_override:
        if output_override.suffix.lower() != ".bin":
            return output_override.with_suffix(".bin")
        return output_override
    return xlsx_path.with_name(f"{xlsx_path.stem}_rebuilt.bin")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Rebuild binary language file from decoded Excel sheet."
    )
    parser.add_argument(
        "xlsx_file",
        nargs="?",
        help="Path to .xlsx produced/edited from decode output. If omitted, file picker opens.",
    )
    parser.add_argument("-o", "--output", help="Output .bin file path")
    parser.add_argument(
        "--template",
        help="Optional original .bin file to copy first 4 bytes from (recommended).",
    )
    parser.add_argument(
        "--encoding",
        default="utf-8",
        help="Default text encoding used when Note column has no detected encoding (default: utf-8).",
    )
    parser.add_argument(
        "--entry-size",
        type=int,
        choices=[4, 8],
        default=4,
        help="Pointer table entry size (4 for contiguous pointers, 8 for pointer+00000000).",
    )
    parser.add_argument(
        "--mode",
        choices=["addressed", "inplace", "repack"],
        default="addressed",
        help="addressed: preserve StringOffsetHex pointers exactly; inplace: patch template in place; repack: rebuild pointers and string block.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail if text cannot be encoded with selected encoding instead of replacing chars.",
    )
    parser.add_argument(
        "--header-word",
        help="Override first 4 bytes as little-endian integer (example: 0x1C21 or 7201).",
    )

    args = parser.parse_args()

    if args.xlsx_file:
        xlsx_path = Path(args.xlsx_file)
    else:
        selected = choose_file_with_dialog(
            title="Select decoded Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if selected is None:
            print("No Excel file selected. Aborting.")
            return 1
        xlsx_path = selected

    if not xlsx_path.exists() or not xlsx_path.is_file():
        print(f"Excel file not found: {xlsx_path}")
        return 1

    force_header_word: Optional[int] = None
    if args.header_word:
        word_text = str(args.header_word).strip().lower()
        if word_text.startswith("0x"):
            force_header_word = int(word_text, 16)
        else:
            force_header_word = int(word_text)

    wb = load_workbook(xlsx_path, data_only=True)
    if "DecodedStrings" not in wb.sheetnames:
        print("Sheet 'DecodedStrings' was not found in workbook.")
        return 1

    sheet = wb["DecodedStrings"]
    header_row = find_header_row(sheet)
    col_map = get_col_index_map(sheet, header_row)
    rows = build_rows(sheet, col_map, header_row)

    if not rows:
        print("No rows found to rebuild.")
        return 1

    source_hint = extract_source_file_from_sheet(sheet)
    template = Path(args.template) if args.template else None
    header_prefix = choose_header_prefix(template, source_hint, force_header_word)

    warnings: List[str] = []
    if args.mode == "inplace":
        template_for_inplace: Optional[Path] = None
        if template and template.exists() and template.is_file():
            template_for_inplace = template
        elif source_hint and source_hint.exists() and source_hint.is_file():
            template_for_inplace = source_hint

        if template_for_inplace is None:
            print(
                "In-place mode requires a template binary. Provide --template or ensure Source file path exists."
            )
            return 1

        template_bytes = template_for_inplace.read_bytes()
        rebuilt, inplace_warnings = build_inplace_from_template(
            template_bytes=template_bytes,
            rows=rows,
            default_encoding=args.encoding,
            strict=args.strict,
        )
        warnings.extend(inplace_warnings)
    elif args.mode == "addressed":
        fixed_size: Optional[int] = None
        for candidate in [template, source_hint]:
            if candidate and candidate.exists() and candidate.is_file():
                fixed_size = candidate.stat().st_size
                break

        rebuilt, warnings = build_addressed_binary(
            rows=rows,
            header_prefix=header_prefix,
            entry_size=args.entry_size,
            default_encoding=args.encoding,
            strict=args.strict,
            fixed_size=fixed_size,
        )
    else:
        encoded_rows, warnings = encode_text_rows(
            rows=rows,
            default_encoding=args.encoding,
            strict=args.strict,
        )

        rebuilt = build_binary(
            encoded_rows=encoded_rows,
            header_prefix=header_prefix,
            entry_size=args.entry_size,
        )

    output_path = build_output_path(xlsx_path, Path(args.output) if args.output else None)
    output_path.write_bytes(rebuilt)

    print(f"Rebuilt {len(rows)} strings.")
    print(f"Output written to: {output_path}")
    print(f"Binary size: {len(rebuilt)} bytes")
    if args.mode in ("repack", "addressed"):
        print(f"Entry size: {args.entry_size}")
        print(f"Header first word: 0x{int.from_bytes(header_prefix, 'little'):08X}")
        if args.mode == "addressed":
            print("Mode: addressed (pointer values copied from StringOffsetHex)")
    else:
        print("Mode: inplace (pointer table/layout preserved from template)")
    if warnings:
        print(f"Warnings: {len(warnings)}")
        for msg in warnings[:10]:
            print(f"  - {msg}")
        if len(warnings) > 10:
            print(f"  ... and {len(warnings) - 10} more")

    return 0


if __name__ == "__main__":
    sys.exit(main())
